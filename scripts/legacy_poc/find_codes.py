#!/usr/bin/env python3
import argparse
import re
from openpyxl import load_workbook


CODE_RE = re.compile(r'[A-Za-z]{2}\d{2}-[A-Za-z0-9\-/]+')


def find_codes(path, out):
    wb = load_workbook(filename=path, keep_vba=True, data_only=True)
    hits = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        max_r = ws.max_row
        max_c = ws.max_column
        for r in range(1, max_r+1):
            for c in range(1, max_c+1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is None:
                    continue
                if not isinstance(val, str):
                    try:
                        val = str(val)
                    except Exception:
                        continue
                for m in CODE_RE.findall(val):
                    hits.append((sheet, r, c, m, val.strip()))
    # write markdown
    with open(out, 'w', encoding='utf-8') as f:
        f.write('# Codes found in workbook\n\n')
        if not hits:
            f.write('No codes found matching pattern.\n')
            return 0
        by_sheet = {}
        for s,r,c,m,raw in hits:
            by_sheet.setdefault(s, []).append((r,c,m,raw))
        for s in by_sheet:
            f.write(f'## Sheet: {s}  \n')
            for r,c,m,raw in by_sheet[s]:
                f.write(f'- Row {r}, Col {c}: `{m}` (cell contains: "{raw}")\n')
            f.write('\n')
    return 0


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('file')
    parser.add_argument('--out', '-o', default='docs/codes_found.md')
    args = parser.parse_args()
    rc = find_codes(args.file, args.out)
    if rc == 0:
        print('Wrote', args.out)


if __name__ == '__main__':
    main()
