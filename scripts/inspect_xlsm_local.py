#!/usr/bin/env python3
"""Inspeciona um arquivo .xlsm e gera um inventário simples.

Uso:
  python scripts/inspect_xlsm_local.py "Novo_template_BITin_V2 TESTE.xlsm" --out docs/inventory.md
"""
import argparse
from openpyxl import load_workbook


def inspect(path):
    wb = load_workbook(filename=path, keep_vba=True, data_only=True)
    data = []
    data.append(f"File: {path}\n")
    data.append(f"Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}\n")
    # defined names
    try:
        defined = list(wb.defined_names.keys())
    except Exception:
        defined = []
    data.append(f"Defined names ({len(defined)}): {', '.join(defined)}\n")

    # count formulas per sheet
    for name in wb.sheetnames:
        ws = wb[name]
        formula_count = 0
        cell_count = 0
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                cell_count += 1
                if cell.data_type == 'f' or (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                    formula_count += 1
        data.append(f"Sheet '{name}': cells={cell_count}, formulas={formula_count}\n")

    return '\n'.join(data)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('file')
    parser.add_argument('--out', '-o', default=None)
    args = parser.parse_args()
    report = inspect(args.file)
    if args.out:
        with open(args.out, 'w', encoding='utf-8') as f:
            f.write(report)
        print(f'Report written to {args.out}')
    else:
        print(report)


if __name__ == '__main__':
    main()
